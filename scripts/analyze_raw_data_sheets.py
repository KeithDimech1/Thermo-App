#!/usr/bin/env python3
"""
Analyze NAT and SEROLOGY raw data sheets
to prepare for database import
"""

from openpyxl import load_workbook
import json

file_path = 'build-data/documentation/Copy of 1. Summary Tables CV Events.xlsx'

# Load workbook
wb = load_workbook(file_path, data_only=True)

print('='*80)
print('ANALYZING RAW DATA SHEETS FOR DATABASE IMPORT')
print('='*80)

# Analyze NAT raw data
print('\n' + '='*80)
print('=== NAT RAW DATA ===')
print('='*80)
nat_sheet = wb['NAT raw data']
print(f'Dimensions: {nat_sheet.max_row} rows x {nat_sheet.max_column} columns')
print(f'Data rows (excluding header): {nat_sheet.max_row - 1}')
print()

print('First 3 rows (to understand structure):')
for row_idx in range(1, min(4, nat_sheet.max_row + 1)):
    print(f'  Row {row_idx}:', end='')
    for col_idx in range(1, min(6, nat_sheet.max_column + 1)):
        value = nat_sheet.cell(row_idx, col_idx).value
        print(f' | {value}', end='')
    print()
print()

# Headers appear to be in row 3
print('Column headers (from row 3):')
nat_headers = []
for col in range(1, nat_sheet.max_column + 1):
    header = nat_sheet.cell(3, col).value
    nat_headers.append(header)
    print(f'  {col:2d}. {header}')
print()

print('Sample first 3 data rows (starting row 4):')
for row_idx in range(4, min(7, nat_sheet.max_row + 1)):
    print(f'\n  Row {row_idx}:')
    for col_idx, header in enumerate(nat_headers, 1):
        value = nat_sheet.cell(row_idx, col_idx).value
        if value:
            print(f'    {header}: {value}')

# Analyze SEROLOGY raw data
print('\n' + '='*80)
print('=== SEROLOGY RAW DATA ===')
print('='*80)
ser_sheet = wb['SEROLOGY raw data']
print(f'Dimensions: {ser_sheet.max_row} rows x {ser_sheet.max_column} columns')
print(f'Data rows (excluding header): {ser_sheet.max_row - 1}')
print()

print('First 3 rows (to understand structure):')
for row_idx in range(1, min(4, ser_sheet.max_row + 1)):
    print(f'  Row {row_idx}:', end='')
    for col_idx in range(1, min(6, ser_sheet.max_column + 1)):
        value = ser_sheet.cell(row_idx, col_idx).value
        print(f' | {value}', end='')
    print()
print()

# Headers appear to be in row 3
print('Column headers (from row 3):')
ser_headers = []
for col in range(1, ser_sheet.max_column + 1):
    header = ser_sheet.cell(3, col).value
    ser_headers.append(header)
    print(f'  {col:2d}. {header}')
print()

print('Sample first 3 data rows (starting row 4):')
for row_idx in range(4, min(7, ser_sheet.max_row + 1)):
    print(f'\n  Row {row_idx}:')
    for col_idx, header in enumerate(ser_headers, 1):
        value = ser_sheet.cell(row_idx, col_idx).value
        if value:
            print(f'    {header}: {value}')

# Analyze SEROLOGY to include (current data)
print('\n' + '='*80)
print('=== SEROLOGY TO INCLUDE (CURRENT IN DB) ===')
print('='*80)
inc_sheet = wb['SEROLOGY to include']
print(f'Dimensions: {inc_sheet.max_row} rows x {inc_sheet.max_column} columns')
print(f'Data rows (excluding header): {inc_sheet.max_row - 1}')
print()

print('Column headers:')
inc_headers = []
for col in range(1, inc_sheet.max_column + 1):
    header = inc_sheet.cell(1, col).value
    inc_headers.append(header)
    print(f'  {col:2d}. {header}')
print()

# Compare headers
print('\n' + '='*80)
print('HEADER COMPARISON')
print('='*80)
print(f'\nNAT headers: {len(nat_headers)} columns')
print(nat_headers)
print(f'\nSEROLOGY raw headers: {len(ser_headers)} columns')
print(ser_headers)
print(f'\nSEROLOGY to include headers: {len(inc_headers)} columns')
print(inc_headers)

# Check for marker column
print('\n' + '='*80)
print('KEY DIFFERENCES')
print('='*80)
if 'Marker' in inc_headers and 'Marker' not in ser_headers:
    print('✓ SEROLOGY to include has "Marker" column (column 0)')
    print('✗ SEROLOGY raw data does NOT have "Marker" column')
    print('  → Will need to infer marker from other columns')
print()

if len(inc_headers) > len(ser_headers):
    print(f'✓ SEROLOGY to include has {len(inc_headers) - len(ser_headers)} extra columns')
    extra_cols = set(inc_headers) - set(ser_headers)
    print(f'  Extra columns: {extra_cols}')

# Collect unique values for key columns
print('\n' + '='*80)
print('UNIQUE VALUES IN KEY COLUMNS')
print('='*80)

# QC Sample in NAT (data starts row 4)
nat_samples = set()
if 'Sample' in nat_headers:
    qc_col = nat_headers.index('Sample') + 1
    for row in range(4, nat_sheet.max_row + 1):
        value = nat_sheet.cell(row, qc_col).value
        if value:
            nat_samples.add(value)
    print(f'\nNAT - Unique QC Samples ({len(nat_samples)}):')
    for sample in sorted(nat_samples):
        print(f'  - {sample}')

# QC Sample in SEROLOGY raw (data starts row 4)
ser_samples = set()
if 'Sample' in ser_headers:
    qc_col = ser_headers.index('Sample') + 1
    for row in range(4, ser_sheet.max_row + 1):
        value = ser_sheet.cell(row, qc_col).value
        if value:
            ser_samples.add(value)
    print(f'\nSEROLOGY raw - Unique QC Samples ({len(ser_samples)}):')
    for sample in sorted(ser_samples):
        print(f'  - {sample}')

# Assay in NAT
nat_assays = set()
if 'Assay' in nat_headers:
    assay_col = nat_headers.index('Assay') + 1
    for row in range(4, nat_sheet.max_row + 1):
        value = nat_sheet.cell(row, assay_col).value
        if value:
            nat_assays.add(value)
    print(f'\nNAT - Unique Assays ({len(nat_assays)}):')
    for assay in sorted(nat_assays)[:10]:  # Show first 10
        print(f'  - {assay}')
    if len(nat_assays) > 10:
        print(f'  ... and {len(nat_assays) - 10} more')

# Assay in SEROLOGY raw
ser_assays = set()
if 'Assay' in ser_headers:
    assay_col = ser_headers.index('Assay') + 1
    for row in range(4, ser_sheet.max_row + 1):
        value = ser_sheet.cell(row, assay_col).value
        if value:
            ser_assays.add(value)
    print(f'\nSEROLOGY raw - Unique Assays ({len(ser_assays)}):')
    for assay in sorted(ser_assays)[:10]:  # Show first 10
        print(f'  - {assay}')
    if len(ser_assays) > 10:
        print(f'  ... and {len(ser_assays) - 10} more')

wb.close()

print('\n' + '='*80)
print('ANALYSIS COMPLETE')
print('='*80)
