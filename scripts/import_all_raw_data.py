#!/usr/bin/env python3
"""
Complete Raw Data Import - NAT + SEROLOGY Extended
Handles normalized schema with foreign keys
"""

import os
import sys
from openpyxl import load_workbook
import psycopg2
from dotenv import load_dotenv

# Load environment variables
load_dotenv('.env.local')

DATABASE_URL = os.getenv('DATABASE_URL')
if not DATABASE_URL:
    print('Error: DATABASE_URL not found')
    sys.exit(1)

EXCEL_PATH = 'build-data/documentation/Copy of 1. Summary Tables CV Events.xlsx'

# ============================================================================
# DATABASE HELPERS
# ============================================================================

def get_or_create_category(conn, name):
    """Get or create category by name"""
    cursor = conn.cursor()

    # Try to find existing
    cursor.execute("SELECT id FROM categories WHERE name = %s", (name,))
    result = cursor.fetchone()
    if result:
        cursor.close()
        return result[0]

    # Create new
    cursor.execute(
        "INSERT INTO categories (name) VALUES (%s) ON CONFLICT (name) DO NOTHING RETURNING id",
        (name,)
    )
    result = cursor.fetchone()
    if result:
        category_id = result[0]
    else:
        # Race condition - another process created it
        cursor.execute("SELECT id FROM categories WHERE name = %s", (name,))
        category_id = cursor.fetchone()[0]
    cursor.close()
    return category_id

def get_or_create_pathogen(conn, name, category_name=None):
    """Get or create pathogen by name"""
    cursor = conn.cursor()

    # Try to find existing
    cursor.execute("SELECT id FROM pathogens WHERE name = %s", (name,))
    result = cursor.fetchone()
    if result:
        cursor.close()
        return result[0]

    # Get category ID if provided
    category_id = None
    if category_name:
        category_id = get_or_create_category(conn, category_name)

    # Create new
    cursor.execute(
        "INSERT INTO pathogens (name, category_id) VALUES (%s, %s) RETURNING id",
        (name, category_id)
    )
    pathogen_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    return pathogen_id

def get_or_create_marker(conn, name, pathogen_name=None, category_name=None, antibody_type=None):
    """Get or create marker by name"""
    cursor = conn.cursor()

    # Try to find existing
    cursor.execute("SELECT id FROM markers WHERE name = %s", (name,))
    result = cursor.fetchone()
    if result:
        cursor.close()
        return result[0]

    # Get pathogen and category IDs
    pathogen_id = None
    category_id = None

    if pathogen_name:
        pathogen_id = get_or_create_pathogen(conn, pathogen_name, category_name)

    if category_name and not category_id:
        category_id = get_or_create_category(conn, category_name)

    # Create new marker
    cursor.execute(
        """INSERT INTO markers (name, pathogen_id, category_id, antibody_type, marker_type)
           VALUES (%s, %s, %s, %s, %s) RETURNING id""",
        (name, pathogen_id, category_id, antibody_type, 'Antibody' if antibody_type else 'Molecular')
    )
    marker_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    return marker_id

def get_or_create_manufacturer(conn, name):
    """Get or create manufacturer by name"""
    cursor = conn.cursor()

    # Try to find existing
    cursor.execute("SELECT id FROM manufacturers WHERE name = %s", (name,))
    result = cursor.fetchone()
    if result:
        cursor.close()
        return result[0]

    # Create new
    cursor.execute(
        "INSERT INTO manufacturers (name, total_assays) VALUES (%s, 0) RETURNING id",
        (name,)
    )
    manufacturer_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    return manufacturer_id

def get_or_create_assay(conn, name, manufacturer_name, platform=None, methodology=None):
    """Get or create assay by name and manufacturer"""
    cursor = conn.cursor()

    # Get manufacturer ID
    manufacturer_id = get_or_create_manufacturer(conn, manufacturer_name)

    # Try to find existing
    cursor.execute(
        "SELECT id FROM assays WHERE name = %s AND manufacturer_id = %s",
        (name, manufacturer_id)
    )
    result = cursor.fetchone()
    if result:
        cursor.close()
        return result[0]

    # Create new
    cursor.execute(
        """INSERT INTO assays (name, manufacturer_id, platform, methodology)
           VALUES (%s, %s, %s, %s) RETURNING id""",
        (name, manufacturer_id, platform, methodology)
    )
    assay_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    return assay_id

def get_or_create_qc_sample(conn, name, matrix='Serum'):
    """Get or create QC sample by name"""
    cursor = conn.cursor()

    # Try to find existing
    cursor.execute("SELECT id FROM qc_samples WHERE name = %s", (name,))
    result = cursor.fetchone()
    if result:
        cursor.close()
        return result[0]

    # Create new
    cursor.execute(
        "INSERT INTO qc_samples (name, matrix) VALUES (%s, %s) RETURNING id",
        (name, matrix)
    )
    qc_sample_id = cursor.fetchone()[0]
    conn.commit()
    cursor.close()
    return qc_sample_id

def config_exists(conn, marker_id, assay_id, qc_sample_id):
    """Check if test configuration already exists"""
    cursor = conn.cursor()
    cursor.execute(
        """SELECT id FROM test_configurations
           WHERE marker_id = %s AND assay_id = %s AND qc_sample_id = %s""",
        (marker_id, assay_id, qc_sample_id)
    )
    result = cursor.fetchone()
    cursor.close()
    return result is not None

# ============================================================================
# DATA PARSING HELPERS
# ============================================================================

def extract_manufacturer(assay_name):
    """Extract manufacturer from assay name"""
    assay_upper = assay_name.upper()

    manufacturers = {
        'ABBOTT': 'Abbott',
        'ROCHE': 'Roche',
        'SIEMENS': 'Siemens',
        'DIASORIN': 'DiaSorin',
        'BIOMERIEUX': 'bioMerieux',
        'BECKMAN': 'Beckman Coulter',
        'ORTHO': 'Ortho Clinical Diagnostics',
        'QIAGEN': 'QIAGEN',
        'GRIFOLS': 'Grifols',
        'WERFEN': 'Werfen',
        'SNIBE': 'SNIBE',
        'EUROIMMUN': 'Euroimmun',
        'LIAISON': 'DiaSorin',
        'VIDAS': 'bioMerieux',
        'ARCHITECT': 'Abbott',
        'ALINITY': 'Abbott',
        'COBAS': 'Roche',
        'ELECSYS': 'Roche',
        'ADVIA': 'Siemens',
        'CENTAUR': 'Siemens',
        'IMMULITE': 'Siemens',
        'VITROS': 'Ortho Clinical Diagnostics',
        'AUSDIAGNOSTICS': 'AusDiagnostics',
        'CEPHEID': 'Cepheid',
        'HOLOGIC': 'Hologic',
    }

    for key, manufacturer in manufacturers.items():
        if key in assay_upper:
            return manufacturer

    return 'Unknown'

def parse_nat_marker(sample_name, assay_name):
    """Parse NAT marker from sample and assay names"""
    combined = (sample_name + ' ' + assay_name).upper()

    markers = {
        'CMV': ('CMV', 'Cytomegalovirus (CMV)', 'Viral', None),
        'HCV': ('HCV', 'Hepatitis C Virus (HCV)', 'Viral', None),
        'HBV': ('HBV', 'Hepatitis B Virus (HBV)', 'Viral', None),
        'HIV': ('HIV', 'Human Immunodeficiency Virus (HIV)', 'Viral', None),
        'EBV': ('EBV', 'Epstein-Barr Virus (EBV)', 'Viral', None),
        'HPV': ('HPV', 'Human Papillomavirus (HPV)', 'Viral', None),
        'COVID': ('SARS-CoV-2', 'SARS-CoV-2', 'Viral', None),
        'SARS': ('SARS-CoV-2', 'SARS-CoV-2', 'Viral', None),
        'CT': ('Chlamydia trachomatis', 'Chlamydia trachomatis', 'Bacterial', None),
        'NG': ('Neisseria gonorrhoeae', 'Neisseria gonorrhoeae', 'Bacterial', None),
    }

    for key, (marker, pathogen, category, antibody) in markers.items():
        if key in combined:
            return marker, pathogen, category, antibody

    return 'Unknown NAT', 'Unknown', 'Molecular', None

def parse_serology_marker(assay_name):
    """Parse serology marker from assay name"""
    assay_upper = assay_name.upper()

    # IgG markers
    if 'CMV' in assay_upper and 'IGG' in assay_upper:
        return 'anti-CMV IgG', 'Cytomegalovirus (CMV)', 'TORCH', 'IgG'
    elif 'EBV' in assay_upper and 'IGG' in assay_upper:
        return 'anti-EBV IgG', 'Epstein-Barr Virus (EBV)', 'Viral', 'IgG'
    elif 'TOXO' in assay_upper and 'IGG' in assay_upper:
        return 'anti-Toxoplasma IgG', 'Toxoplasma gondii', 'TORCH', 'IgG'
    elif 'RUBELLA' in assay_upper and 'IGG' in assay_upper:
        return 'anti-Rubella IgG', 'Rubella virus', 'TORCH', 'IgG'

    # IgM markers
    elif 'CMV' in assay_upper and 'IGM' in assay_upper:
        return 'anti-CMV IgM', 'Cytomegalovirus (CMV)', 'TORCH', 'IgM'
    elif 'EBV' in assay_upper and 'IGM' in assay_upper:
        return 'anti-EBV IgM', 'Epstein-Barr Virus (EBV)', 'Viral', 'IgM'
    elif 'TOXO' in assay_upper and 'IGM' in assay_upper:
        return 'anti-Toxoplasma IgM', 'Toxoplasma gondii', 'TORCH', 'IgM'
    elif 'RUBELLA' in assay_upper and 'IGM' in assay_upper:
        return 'anti-Rubella IgM', 'Rubella virus', 'TORCH', 'IgM'

    # HCV
    elif 'HCV' in assay_upper:
        if 'IGG' in assay_upper:
            return 'anti-HCV IgG', 'Hepatitis C Virus (HCV)', 'Viral', 'IgG'
        else:
            return 'anti-HCV', 'Hepatitis C Virus (HCV)', 'Viral', None

    # HAV
    elif 'HAV' in assay_upper:
        if 'IGM' in assay_upper:
            return 'anti-HAV IgM', 'Hepatitis A Virus (HAV)', 'Viral', 'IgM'
        elif 'IGG' in assay_upper:
            return 'anti-HAV IgG', 'Hepatitis A Virus (HAV)', 'Viral', 'IgG'
        else:
            return 'anti-HAV', 'Hepatitis A Virus (HAV)', 'Viral', None

    # HBV markers
    elif 'HBSAG' in assay_upper:
        return 'HBsAg', 'Hepatitis B Virus (HBV)', 'Viral', None
    elif 'HBS' in assay_upper:
        return 'anti-HBs', 'Hepatitis B Virus (HBV)', 'Viral', 'IgG'
    elif 'HBC' in assay_upper and 'IGM' in assay_upper:
        return 'anti-HBc IgM', 'Hepatitis B Virus (HBV)', 'Viral', 'IgM'
    elif 'HBC' in assay_upper:
        return 'anti-HBc', 'Hepatitis B Virus (HBV)', 'Viral', None
    elif 'HBE' in assay_upper:
        return 'anti-HBe', 'Hepatitis B Virus (HBV)', 'Viral', None

    # HIV
    elif 'HIV' in assay_upper:
        if 'P24' in assay_upper:
            return 'HIV p24 antigen', 'Human Immunodeficiency Virus (HIV)', 'Viral', None
        elif 'HIV-2' in assay_upper or 'HIV2' in assay_upper:
            return 'anti-HIV-2', 'Human Immunodeficiency Virus (HIV)', 'Viral', None
        else:
            return 'anti-HIV-1+2', 'Human Immunodeficiency Virus (HIV)', 'Viral', None

    # Syphilis
    elif 'SYPHILIS' in assay_upper or 'TP' in assay_upper and 'AB' in assay_upper:
        return 'anti-Treponema pallidum', 'Treponema pallidum', 'Bacterial', None

    return 'Unknown', 'Unknown', 'Unknown', None

def calculate_quality_rating(cv_lt_10_pct):
    """Calculate quality rating from CV <10% percentage"""
    if cv_lt_10_pct >= 0.95:
        return 'excellent'
    elif cv_lt_10_pct >= 0.85:
        return 'good'
    elif cv_lt_10_pct >= 0.70:
        return 'acceptable'
    else:
        return 'poor'

# ============================================================================
# IMPORT FUNCTIONS
# ============================================================================

def import_nat_data(conn):
    """Import NAT raw data"""
    print('\n' + '='*80)
    print('IMPORTING NAT DATA')
    print('='*80)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['NAT raw data']

    imported = 0
    skipped = 0
    errors = 0

    for row_idx in range(4, ws.max_row + 1):
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events_examined = ws.cell(row_idx, 3).value
        cv_lt_10_pct = ws.cell(row_idx, 5).value or 0
        cv_10_15_pct = ws.cell(row_idx, 7).value or 0
        cv_15_20_pct = ws.cell(row_idx, 9).value or 0
        cv_gt_20_pct = ws.cell(row_idx, 11).value or 0

        if not assay_name or not sample_name or assay_name == 'Assay':
            continue

        try:
            # Parse marker
            marker_name, pathogen_name, category_name, antibody_type = parse_nat_marker(sample_name, assay_name)

            # Get or create entities
            marker_id = get_or_create_marker(conn, marker_name, pathogen_name, category_name, antibody_type)
            manufacturer_name = extract_manufacturer(assay_name)
            assay_id = get_or_create_assay(conn, assay_name, manufacturer_name, platform=None, methodology='PCR')
            qc_sample_id = get_or_create_qc_sample(conn, sample_name, matrix='Plasma')

            # Check if exists
            if config_exists(conn, marker_id, assay_id, qc_sample_id):
                skipped += 1
                continue

            # Calculate quality
            quality_rating = calculate_quality_rating(cv_lt_10_pct)

            # Insert test configuration
            cursor = conn.cursor()
            cursor.execute(
                """INSERT INTO test_configurations (
                    marker_id, assay_id, qc_sample_id, test_type,
                    events_examined, quality_rating, inclusion_group
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (marker_id, assay_id, qc_sample_id, 'nat',
                 int(events_examined) if events_examined else 0,
                 quality_rating, 'nat_data')
            )

            # Insert CV measurements
            cursor.execute(
                """INSERT INTO cv_measurements (
                    test_config_id, cv_category, percentage
                ) VALUES
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), 'lt_10', %s),
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), '10_15', %s),
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), '15_20', %s),
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), 'gt_20', %s)
                """,
                (marker_id, assay_id, qc_sample_id, float(cv_lt_10_pct) * 100,
                 marker_id, assay_id, qc_sample_id, float(cv_10_15_pct) * 100,
                 marker_id, assay_id, qc_sample_id, float(cv_15_20_pct) * 100,
                 marker_id, assay_id, qc_sample_id, float(cv_gt_20_pct) * 100)
            )

            conn.commit()
            cursor.close()
            imported += 1

            if imported % 10 == 0:
                print(f'  Progress: {imported} rows imported...', end='\r')

        except Exception as e:
            print(f'\n  Error on row {row_idx}: {e}')
            conn.rollback()
            errors += 1

    wb.close()
    print(f'\n✓ NAT import complete: {imported} imported, {skipped} skipped, {errors} errors')
    return imported

def import_serology_extended_data(conn):
    """Import extended SEROLOGY raw data"""
    print('\n' + '='*80)
    print('IMPORTING SEROLOGY EXTENDED DATA')
    print('='*80)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SEROLOGY raw data']

    imported = 0
    skipped = 0
    errors = 0

    for row_idx in range(4, ws.max_row + 1):
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events_examined = ws.cell(row_idx, 3).value
        cv_lt_10_pct = ws.cell(row_idx, 5).value or 0
        cv_10_15_pct = ws.cell(row_idx, 7).value or 0
        cv_15_20_pct = ws.cell(row_idx, 9).value or 0
        cv_gt_20_pct = ws.cell(row_idx, 11).value or 0

        if not assay_name or not sample_name or assay_name == 'Assay':
            continue

        try:
            # Parse marker
            marker_name, pathogen_name, category_name, antibody_type = parse_serology_marker(assay_name)

            # Get or create entities
            marker_id = get_or_create_marker(conn, marker_name, pathogen_name, category_name, antibody_type)
            manufacturer_name = extract_manufacturer(assay_name)
            assay_id = get_or_create_assay(conn, assay_name, manufacturer_name, platform=None, methodology=None)
            qc_sample_id = get_or_create_qc_sample(conn, sample_name, matrix='Serum')

            # Check if exists
            if config_exists(conn, marker_id, assay_id, qc_sample_id):
                skipped += 1
                continue

            # Calculate quality
            quality_rating = calculate_quality_rating(cv_lt_10_pct)

            # Insert test configuration
            cursor = conn.cursor()
            cursor.execute(
                """INSERT INTO test_configurations (
                    marker_id, assay_id, qc_sample_id, test_type,
                    events_examined, quality_rating, inclusion_group
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (marker_id, assay_id, qc_sample_id, 'serology',
                 int(events_examined) if events_examined else 0,
                 quality_rating, 'serology_extended')
            )

            # Insert CV measurements
            cursor.execute(
                """INSERT INTO cv_measurements (
                    test_config_id, cv_category, percentage
                ) VALUES
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), 'lt_10', %s),
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), '10_15', %s),
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), '15_20', %s),
                    ((SELECT id FROM test_configurations WHERE marker_id=%s AND assay_id=%s AND qc_sample_id=%s), 'gt_20', %s)
                """,
                (marker_id, assay_id, qc_sample_id, float(cv_lt_10_pct) * 100,
                 marker_id, assay_id, qc_sample_id, float(cv_10_15_pct) * 100,
                 marker_id, assay_id, qc_sample_id, float(cv_15_20_pct) * 100,
                 marker_id, assay_id, qc_sample_id, float(cv_gt_20_pct) * 100)
            )

            conn.commit()
            cursor.close()
            imported += 1

            if imported % 10 == 0:
                print(f'  Progress: {imported} rows imported...', end='\r')

        except Exception as e:
            print(f'\n  Error on row {row_idx}: {e}')
            conn.rollback()
            errors += 1

    wb.close()
    print(f'\n✓ SEROLOGY extended import complete: {imported} imported, {skipped} skipped, {errors} errors')
    return imported

# ============================================================================
# MAIN
# ============================================================================

def main():
    print('='*80)
    print('RAW DATA IMPORT - COMPLETE DATASET')
    print('='*80)
    print(f'\nDatabase: {DATABASE_URL[:60]}...')
    print(f'Excel: {EXCEL_PATH}')

    try:
        conn = psycopg2.connect(DATABASE_URL)

        # Import NAT data
        nat_count = import_nat_data(conn)

        # Import SEROLOGY extended data
        serology_count = import_serology_extended_data(conn)

        # Final summary
        print('\n' + '='*80)
        print('IMPORT COMPLETE')
        print('='*80)

        cursor = conn.cursor()
        cursor.execute("""
            SELECT inclusion_group, test_type, COUNT(*) as count
            FROM test_configurations
            GROUP BY inclusion_group, test_type
            ORDER BY inclusion_group, test_type
        """)

        print('\nFinal database counts:')
        total = 0
        for row in cursor.fetchall():
            print(f'  {row[0]:25s} | {row[1]:10s} | {row[2]:4d} configs')
            total += row[2]

        print(f'\n  TOTAL: {total} test configurations')

        cursor.close()
        conn.close()

    except Exception as e:
        print(f'\n❌ Error: {e}')
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
