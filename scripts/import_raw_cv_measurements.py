#!/usr/bin/env python3
"""
Import CV Measurements for Raw Data (NAT and Serology Extended)

This script imports the missing cv_measurements records for the NAT and
serology_extended test configurations that were imported but didn't get
their cv_measurements created.

The import_raw_fast.py script only created test_configurations, but the
raw CSV files contain complete CV measurement data that needs to be imported.
"""

import os
import sys
from openpyxl import load_workbook
import psycopg2
from dotenv import load_dotenv
from decimal import Decimal

load_dotenv('.env.local')

DATABASE_URL = os.getenv('DATABASE_URL')
EXCEL_PATH = 'build-data/documentation/Copy of 1. Summary Tables CV Events.xlsx'


def get_test_config_id(conn, marker_name, assay_name, qc_sample_name):
    """Get test_config_id by looking up marker, assay, and qc_sample"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT tc.id
        FROM test_configurations tc
        JOIN markers m ON tc.marker_id = m.id
        JOIN assays a ON tc.assay_id = a.id
        JOIN qc_samples qs ON tc.qc_sample_id = qs.id
        WHERE m.name = %s
        AND a.name = %s
        AND qs.name = %s
    """, (marker_name, assay_name, qc_sample_name))

    result = cursor.fetchone()
    cursor.close()
    return result[0] if result else None


def parse_percentage(value):
    """Convert percentage value to decimal (0-100 range)"""
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        # If already a decimal (0-1), convert to percentage (0-100)
        if 0 <= value <= 1:
            return Decimal(str(value * 100))
        # If already a percentage (0-100), use as is
        return Decimal(str(value))
    return None


def insert_cv_measurement(conn, test_config_id, cv_data):
    """Insert cv_measurement record"""
    cursor = conn.cursor()

    # Check if measurement already exists
    cursor.execute(
        "SELECT id FROM cv_measurements WHERE test_config_id = %s",
        (test_config_id,)
    )
    if cursor.fetchone():
        cursor.close()
        return False  # Already exists

    # Insert new measurement
    cursor.execute("""
        INSERT INTO cv_measurements (
            test_config_id,
            cv_lt_10_count,
            cv_lt_10_percentage,
            cv_10_15_count,
            cv_10_15_percentage,
            cv_15_20_count,
            cv_15_20_percentage,
            cv_gt_20_count,
            cv_gt_20_percentage
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        test_config_id,
        cv_data['cv_lt_10_count'],
        cv_data['cv_lt_10_pct'],
        cv_data['cv_10_15_count'],
        cv_data['cv_10_15_pct'],
        cv_data['cv_15_20_count'],
        cv_data['cv_15_20_pct'],
        cv_data['cv_gt_20_count'],
        cv_data['cv_gt_20_pct']
    ))

    cursor.close()
    return True  # Successfully inserted


def parse_nat_marker(sample_name, assay_name):
    """Parse NAT marker name - MUST match import_raw_fast.py logic"""
    combined = (sample_name + ' ' + assay_name).upper()
    if 'CMV' in combined:
        return 'CMV'
    elif 'HCV' in combined:
        return 'HCV'
    elif 'HBV' in combined:
        return 'HBV'
    elif 'HIV' in combined:
        return 'HIV'
    elif 'HPV' in combined:
        return 'HPV'
    elif 'COVID' in combined or 'SARS' in combined:
        return 'SARS-CoV-2'
    elif 'CT' in combined and 'CHLAMYDIA' not in combined:
        return 'Chlamydia trachomatis'
    elif 'NG' in combined:
        return 'Neisseria gonorrhoeae'
    return 'Unknown NAT'


def parse_serology_marker(assay_name):
    """Parse serology marker - MUST match import_raw_fast.py logic"""
    au = assay_name.upper()
    if 'CMV' in au and 'IGG' in au:
        return 'anti-CMV IgG'
    elif 'CMV' in au and 'IGM' in au:
        return 'anti-CMV IgM'
    elif 'EBV' in au and 'IGG' in au:
        return 'anti-EBV IgG'
    elif 'EBV' in au and 'IGM' in au:
        return 'anti-EBV IgM'
    elif 'TOXO' in au and 'IGG' in au:
        return 'anti-Toxoplasma IgG'
    elif 'TOXO' in au and 'IGM' in au:
        return 'anti-Toxoplasma IgM'
    elif 'RUBELLA' in au and 'IGG' in au:
        return 'anti-Rubella IgG'
    elif 'RUBELLA' in au and 'IGM' in au:
        return 'anti-Rubella IgM'
    elif 'HCV' in au:
        return 'anti-HCV'
    elif 'HAV' in au and 'IGM' in au:
        return 'anti-HAV IgM'
    elif 'HAV' in au and 'IGG' in au:
        return 'anti-HAV IgG'
    elif 'HAV' in au:
        return 'anti-HAV'
    elif 'HBSAG' in au:
        return 'HBsAg'
    elif 'HBS' in au:
        return 'anti-HBs'
    elif 'HBC' in au and 'IGM' in au:
        return 'anti-HBc IgM'
    elif 'HBC' in au:
        return 'anti-HBc'
    elif 'HBE' in au:
        return 'anti-HBe'
    elif 'HIV' in au and 'P24' in au:
        return 'HIV p24 antigen'
    elif 'HIV' in au and ('HIV-2' in au or 'HIV2' in au):
        return 'anti-HIV-2'
    elif 'HIV' in au:
        return 'anti-HIV-1+2'
    elif 'SYPHILIS' in au or 'TP' in au:
        return 'anti-Treponema pallidum'
    return 'Unknown'


def import_nat_cv_measurements(conn):
    """Import cv_measurements for NAT data"""
    print('\n' + '='*60)
    print('NAT CV MEASUREMENTS IMPORT')
    print('='*60)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['NAT raw data']

    imported = 0
    skipped = 0
    not_found = 0

    for row_idx in range(4, ws.max_row + 1):
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events = ws.cell(row_idx, 3).value

        # CV data from columns 4-11
        cv_lt_10_count = ws.cell(row_idx, 4).value
        cv_lt_10_pct = parse_percentage(ws.cell(row_idx, 5).value)
        cv_gt_10_count = ws.cell(row_idx, 6).value
        cv_gt_10_pct = parse_percentage(ws.cell(row_idx, 7).value)
        cv_gt_15_count = ws.cell(row_idx, 8).value
        cv_gt_15_pct = parse_percentage(ws.cell(row_idx, 9).value)
        cv_gt_20_count = ws.cell(row_idx, 10).value
        cv_gt_20_pct = parse_percentage(ws.cell(row_idx, 11).value)

        if not assay_name or not sample_name or assay_name == 'Assay' or assay_name == 'TOTAL':
            continue

        try:
            # Get marker name
            marker_name = parse_nat_marker(sample_name, assay_name)

            # Get test_config_id
            test_config_id = get_test_config_id(conn, marker_name, assay_name, sample_name)

            if not test_config_id:
                print(f'\n  Warning row {row_idx}: Config not found for {marker_name} / {assay_name[:40]} / {sample_name[:30]}')
                not_found += 1
                continue

            # Calculate 10-15% and 15-20% from the "> threshold" values
            # cv_10_15_pct = cv_gt_10_pct - cv_gt_15_pct
            # cv_15_20_pct = cv_gt_15_pct - cv_gt_20_pct

            cv_10_15_pct = None
            if cv_gt_10_pct is not None and cv_gt_15_pct is not None:
                cv_10_15_pct = cv_gt_10_pct - cv_gt_15_pct

            cv_15_20_pct = None
            if cv_gt_15_pct is not None and cv_gt_20_pct is not None:
                cv_15_20_pct = cv_gt_15_pct - cv_gt_20_pct

            # Prepare cv_data
            cv_data = {
                'cv_lt_10_count': cv_lt_10_count,
                'cv_lt_10_pct': cv_lt_10_pct,
                'cv_10_15_count': None,  # Not directly available in CSV
                'cv_10_15_pct': cv_10_15_pct,
                'cv_15_20_count': None,  # Not directly available in CSV
                'cv_15_20_pct': cv_15_20_pct,
                'cv_gt_20_count': cv_gt_20_count,
                'cv_gt_20_pct': cv_gt_20_pct
            }

            # Insert measurement
            if insert_cv_measurement(conn, test_config_id, cv_data):
                imported += 1
            else:
                skipped += 1

            if (imported + skipped + not_found) % 10 == 0:
                print(f'  Processed: {imported + skipped + not_found} rows...', end='\r')
                conn.commit()  # Commit batch

        except Exception as e:
            print(f'\n  Error row {row_idx}: {e}')
            conn.rollback()

    conn.commit()  # Final commit
    wb.close()
    print(f'\n✓ NAT CV: {imported} imported, {skipped} skipped, {not_found} not found        ')
    return imported


def import_serology_cv_measurements(conn):
    """Import cv_measurements for serology extended data"""
    print('\n' + '='*60)
    print('SEROLOGY EXTENDED CV MEASUREMENTS IMPORT')
    print('='*60)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SEROLOGY raw data']

    imported = 0
    skipped = 0
    not_found = 0

    for row_idx in range(4, ws.max_row + 1):
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events = ws.cell(row_idx, 3).value

        # CV data from columns 4-11
        cv_lt_10_count = ws.cell(row_idx, 4).value
        cv_lt_10_pct = parse_percentage(ws.cell(row_idx, 5).value)
        cv_gt_10_count = ws.cell(row_idx, 6).value
        cv_gt_10_pct = parse_percentage(ws.cell(row_idx, 7).value)
        cv_gt_15_count = ws.cell(row_idx, 8).value
        cv_gt_15_pct = parse_percentage(ws.cell(row_idx, 9).value)
        cv_gt_20_count = ws.cell(row_idx, 10).value
        cv_gt_20_pct = parse_percentage(ws.cell(row_idx, 11).value)

        if not assay_name or not sample_name or assay_name == 'Assay' or assay_name == 'TOTAL':
            continue

        try:
            # Get marker name
            marker_name = parse_serology_marker(assay_name)

            # Get test_config_id
            test_config_id = get_test_config_id(conn, marker_name, assay_name, sample_name)

            if not test_config_id:
                print(f'\n  Warning row {row_idx}: Config not found for {marker_name} / {assay_name[:40]} / {sample_name[:30]}')
                not_found += 1
                continue

            # Calculate 10-15% and 15-20% from the "> threshold" values
            cv_10_15_pct = None
            if cv_gt_10_pct is not None and cv_gt_15_pct is not None:
                cv_10_15_pct = cv_gt_10_pct - cv_gt_15_pct

            cv_15_20_pct = None
            if cv_gt_15_pct is not None and cv_gt_20_pct is not None:
                cv_15_20_pct = cv_gt_15_pct - cv_gt_20_pct

            # Prepare cv_data
            cv_data = {
                'cv_lt_10_count': cv_lt_10_count,
                'cv_lt_10_pct': cv_lt_10_pct,
                'cv_10_15_count': None,
                'cv_10_15_pct': cv_10_15_pct,
                'cv_15_20_count': None,
                'cv_15_20_pct': cv_15_20_pct,
                'cv_gt_20_count': cv_gt_20_count,
                'cv_gt_20_pct': cv_gt_20_pct
            }

            # Insert measurement
            if insert_cv_measurement(conn, test_config_id, cv_data):
                imported += 1
            else:
                skipped += 1

            if (imported + skipped + not_found) % 10 == 0:
                print(f'  Processed: {imported + skipped + not_found} rows...', end='\r')
                conn.commit()  # Commit batch

        except Exception as e:
            print(f'\n  Error row {row_idx}: {e}')
            conn.rollback()

    conn.commit()  # Final commit
    wb.close()
    print(f'\n✓ SEROLOGY CV: {imported} imported, {skipped} skipped, {not_found} not found        ')
    return imported


def main():
    print('='*60)
    print('RAW DATA CV MEASUREMENTS IMPORT')
    print('='*60)

    try:
        conn = psycopg2.connect(DATABASE_URL)

        # Import NAT cv_measurements
        nat_count = import_nat_cv_measurements(conn)

        # Import Serology cv_measurements
        serology_count = import_serology_cv_measurements(conn)

        # Final summary
        print('\n' + '='*60)
        print('FINAL SUMMARY')
        print('='*60)
        print(f'\n✓ NAT CV measurements imported: {nat_count}')
        print(f'✓ Serology CV measurements imported: {serology_count}')
        print(f'✓ Total CV measurements imported: {nat_count + serology_count}')

        # Check database totals
        cursor = conn.cursor()
        cursor.execute("SELECT inclusion_group, COUNT(*) FROM test_configurations GROUP BY inclusion_group ORDER BY inclusion_group")
        print('\nTest configurations by group:')
        for row in cursor.fetchall():
            print(f'  {row[0]:20s} : {row[1]:3d} configs')

        cursor.execute("SELECT COUNT(*) FROM cv_measurements")
        total_cv = cursor.fetchone()[0]
        print(f'\nTotal cv_measurements: {total_cv}')

        cursor.close()
        conn.close()

        print('\n✓ Import complete!')

    except Exception as e:
        print(f'\nERROR: {e}')
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    sys.exit(main())
