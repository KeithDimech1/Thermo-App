#!/usr/bin/env python3
"""
Import NAT and SEROLOGY raw data from Excel file
Handles:
- NAT raw data (77 rows)
- SEROLOGY raw data (274 rows, excluding duplicates from "to include")
- Tracks data source via inclusion_group field
- Detects and handles duplicates
"""

import os
import sys
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv
import re

# Load environment variables
load_dotenv('.env.local')

# Database connection
DATABASE_URL = os.getenv('DATABASE_URL')
if not DATABASE_URL:
    print('Error: DATABASE_URL not found in environment variables')
    sys.exit(1)

# Excel file path
EXCEL_PATH = 'build-data/documentation/Copy of 1. Summary Tables CV Events.xlsx'

def connect_db():
    """Connect to database"""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except Exception as e:
        print(f'Error connecting to database: {e}')
        sys.exit(1)

def parse_marker_from_sample_and_analyte(sample, assay_name):
    """
    Infer marker from QC sample name and assay name

    Examples:
    - Optitrol ToRCHG + "CMV IgG" → anti-CMV IgG
    - QConnect CMVNAT Low → CMV NAT
    """
    # NAT markers
    if 'NAT' in sample.upper() or 'PCR' in assay_name.upper() or 'REALTIME' in assay_name.upper():
        if 'CMV' in sample.upper() or 'CMV' in assay_name.upper():
            return 'CMV', 'NAT'
        elif 'HCV' in sample.upper() or 'HCV' in assay_name.upper():
            return 'HCV', 'NAT'
        elif 'HBV' in sample.upper() or 'HBV' in assay_name.upper():
            return 'HBV', 'NAT'
        elif 'HIV' in sample.upper() or 'HIV' in assay_name.upper():
            return 'HIV', 'NAT'
        elif 'EBV' in sample.upper() or 'EBV' in assay_name.upper():
            return 'EBV', 'NAT'
        elif 'HPV' in sample.upper() or 'HPV' in assay_name.upper():
            return 'HPV', 'NAT'
        elif 'COVID' in sample.upper() or 'SARS' in assay_name.upper():
            return 'SARS-CoV-2', 'NAT'

    # Serology markers - extract from assay name
    assay_upper = assay_name.upper()

    # IgG markers
    if 'CMV IGG' in assay_upper or 'CMV-G' in assay_upper:
        return 'anti-CMV IgG', 'serology'
    elif 'EBV' in assay_upper and 'IGG' in assay_upper:
        return 'anti-EBV IgG', 'serology'
    elif 'TOXO' in assay_upper and 'IGG' in assay_upper:
        return 'anti-Toxoplasma IgG', 'serology'
    elif 'RUBELLA' in assay_upper and 'IGG' in assay_upper:
        return 'anti-Rubella IgG', 'serology'
    elif 'HCV' in assay_upper and 'IGG' in assay_upper:
        return 'anti-HCV IgG', 'serology'
    elif 'HCV' in assay_upper and 'IGM' not in assay_upper:
        return 'anti-HCV', 'serology'

    # IgM markers
    elif 'CMV IGM' in assay_upper or 'CMV-M' in assay_upper:
        return 'anti-CMV IgM', 'serology'
    elif 'EBV' in assay_upper and 'IGM' in assay_upper:
        return 'anti-EBV IgM', 'serology'
    elif 'TOXO' in assay_upper and 'IGM' in assay_upper:
        return 'anti-Toxoplasma IgM', 'serology'
    elif 'RUBELLA' in assay_upper and 'IGM' in assay_upper:
        return 'anti-Rubella IgM', 'serology'

    # HAV
    elif 'HAV' in assay_upper:
        if 'IGM' in assay_upper:
            return 'anti-HAV IgM', 'serology'
        elif 'IGG' in assay_upper:
            return 'anti-HAV IgG', 'serology'
        else:
            return 'anti-HAV', 'serology'

    # HBV markers
    elif 'HBSAG' in assay_upper:
        return 'HBsAg', 'serology'
    elif 'HBS' in assay_upper and 'IGG' in assay_upper:
        return 'anti-HBs IgG', 'serology'
    elif 'HBC' in assay_upper and 'IGM' in assay_upper:
        return 'anti-HBc IgM', 'serology'
    elif 'HBC' in assay_upper:
        return 'anti-HBc', 'serology'
    elif 'HBE' in assay_upper:
        return 'anti-HBe', 'serology'

    # HIV
    elif 'HIV' in assay_upper:
        if 'P24' in assay_upper:
            return 'HIV p24 antigen', 'serology'
        elif 'HIV-2' in assay_upper or 'HIV2' in assay_upper:
            return 'anti-HIV-2', 'serology'
        else:
            return 'anti-HIV-1+2', 'serology'

    # Syphilis
    elif 'SYPHILIS' in assay_upper or 'TP' in assay_upper:
        return 'anti-Treponema pallidum', 'serology'

    # Default
    return None, 'serology'

def extract_manufacturer(assay_name):
    """Extract manufacturer from assay name"""
    assay_upper = assay_name.upper()

    manufacturers = {
        'ABBOTT': 'Abbott',
        'ROCHE': 'Roche',
        'SIEMENS': 'Siemens',
        'DIASORIN': 'DiaSorin',
        'BIOMERIEUX': 'bioMerieux',
        'BECKMAN': 'Beckman Coulter',
        'ORTHO': 'Ortho Clinical Diagnostics',
        'QIAGEN': 'QIAGEN',
        'GRIFOLS': 'Grifols',
        'WERFEN': 'Werfen',
        'SNIBE': 'SNIBE',
        'EUROIMMUN': 'Euroimmun',
        'LIAISON': 'DiaSorin',  # LIAISON is DiaSorin brand
        'VIDAS': 'bioMerieux',  # VIDAS is bioMerieux brand
        'ARCHITECT': 'Abbott',
        'ALINITY': 'Abbott',
        'COBAS': 'Roche',
        'ELECSYS': 'Roche',
        'ADVIA': 'Siemens',
        'CENTAUR': 'Siemens',
        'IMMULITE': 'Siemens',
        'VITROS': 'Ortho Clinical Diagnostics',
    }

    for key, manufacturer in manufacturers.items():
        if key in assay_upper:
            return manufacturer

    return 'Unknown'

def get_existing_configs(conn):
    """Get all existing config identifiers to detect duplicates"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT assay_name, qc_sample_name
        FROM test_configurations
    """)
    existing = {(row[0], row[1]) for row in cursor.fetchall()}
    cursor.close()
    return existing

def calculate_quality_rating(cv_lt_10_pct):
    """Calculate quality rating from CV <10% percentage"""
    if cv_lt_10_pct >= 0.95:
        return 'excellent'
    elif cv_lt_10_pct >= 0.85:
        return 'good'
    elif cv_lt_10_pct >= 0.70:
        return 'acceptable'
    else:
        return 'poor'

def import_nat_raw_data(conn):
    """Import NAT raw data from Excel"""
    print('\n' + '='*80)
    print('IMPORTING NAT RAW DATA')
    print('='*80)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['NAT raw data']

    # Get headers from row 3
    headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]

    # Get existing configs to detect duplicates
    existing_configs = get_existing_configs(conn)

    # Prepare data
    configs_to_insert = []
    skipped = 0

    for row_idx in range(4, ws.max_row + 1):
        # Extract values
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events_examined = ws.cell(row_idx, 3).value
        cv_lt_10 = ws.cell(row_idx, 4).value or 0
        cv_lt_10_pct = ws.cell(row_idx, 5).value or 0
        cv_10_15 = ws.cell(row_idx, 6).value or 0
        cv_10_15_pct = ws.cell(row_idx, 7).value or 0
        cv_15_20 = ws.cell(row_idx, 8).value or 0
        cv_15_20_pct = ws.cell(row_idx, 9).value or 0
        cv_gt_20 = ws.cell(row_idx, 10).value or 0
        cv_gt_20_pct = ws.cell(row_idx, 11).value or 0

        # Skip invalid rows
        if not assay_name or assay_name == 'Assay' or not sample_name:
            continue

        # Check for duplicate
        if (assay_name, sample_name) in existing_configs:
            skipped += 1
            continue

        # Infer marker
        marker_name, test_type = parse_marker_from_sample_and_analyte(sample_name, assay_name)

        # Extract manufacturer
        manufacturer_name = extract_manufacturer(assay_name)

        # Calculate quality rating
        quality_rating = calculate_quality_rating(cv_lt_10_pct)

        config = {
            'assay_name': assay_name,
            'qc_sample_name': sample_name,
            'events_examined': int(events_examined) if events_examined else 0,
            'cv_lt_10_percentage': float(cv_lt_10_pct) * 100 if cv_lt_10_pct else 0,
            'cv_10_15_percentage': float(cv_10_15_pct) * 100 if cv_10_15_pct else 0,
            'cv_15_20_percentage': float(cv_15_20_pct) * 100 if cv_15_20_pct else 0,
            'cv_gt_20_percentage': float(cv_gt_20_pct) * 100 if cv_gt_20_pct else 0,
            'quality_rating': quality_rating,
            'test_type': 'nat',
            'marker_name': marker_name,
            'manufacturer_name': manufacturer_name,
            'inclusion_group': 'nat_data',
        }

        configs_to_insert.append(config)

    wb.close()

    print(f'\nFound {len(configs_to_insert)} NAT configurations to import')
    print(f'Skipped {skipped} duplicates')

    if configs_to_insert:
        # Insert into database
        cursor = conn.cursor()

        insert_query = """
            INSERT INTO test_configurations (
                assay_name, qc_sample_name, events_examined,
                cv_lt_10_percentage, cv_10_15_percentage, cv_15_20_percentage, cv_gt_20_percentage,
                quality_rating, test_type, marker_name, manufacturer_name, inclusion_group
            ) VALUES %s
        """

        values = [(
            c['assay_name'], c['qc_sample_name'], c['events_examined'],
            c['cv_lt_10_percentage'], c['cv_10_15_percentage'], c['cv_15_20_percentage'], c['cv_gt_20_percentage'],
            c['quality_rating'], c['test_type'], c['marker_name'], c['manufacturer_name'], c['inclusion_group']
        ) for c in configs_to_insert]

        execute_values(cursor, insert_query, values)
        conn.commit()
        cursor.close()

        print(f'✓ Successfully imported {len(configs_to_insert)} NAT configurations')

    return len(configs_to_insert)

def import_serology_raw_data(conn):
    """Import SEROLOGY raw data from Excel (excluding duplicates)"""
    print('\n' + '='*80)
    print('IMPORTING SEROLOGY RAW DATA')
    print('='*80)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SEROLOGY raw data']

    # Get headers from row 3
    headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]

    # Get existing configs to detect duplicates
    existing_configs = get_existing_configs(conn)

    # Prepare data
    configs_to_insert = []
    skipped = 0

    for row_idx in range(4, ws.max_row + 1):
        # Extract values
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events_examined = ws.cell(row_idx, 3).value
        cv_lt_10 = ws.cell(row_idx, 4).value or 0
        cv_lt_10_pct = ws.cell(row_idx, 5).value or 0
        cv_10_15 = ws.cell(row_idx, 6).value or 0
        cv_10_15_pct = ws.cell(row_idx, 7).value or 0
        cv_15_20 = ws.cell(row_idx, 8).value or 0
        cv_15_20_pct = ws.cell(row_idx, 9).value or 0
        cv_gt_20 = ws.cell(row_idx, 10).value or 0
        cv_gt_20_pct = ws.cell(row_idx, 11).value or 0

        # Skip invalid rows
        if not assay_name or assay_name == 'Assay' or not sample_name:
            continue

        # Check for duplicate (already in "SEROLOGY to include")
        if (assay_name, sample_name) in existing_configs:
            skipped += 1
            continue

        # Infer marker
        marker_name, test_type = parse_marker_from_sample_and_analyte(sample_name, assay_name)

        # Extract manufacturer
        manufacturer_name = extract_manufacturer(assay_name)

        # Calculate quality rating
        quality_rating = calculate_quality_rating(cv_lt_10_pct)

        config = {
            'assay_name': assay_name,
            'qc_sample_name': sample_name,
            'events_examined': int(events_examined) if events_examined else 0,
            'cv_lt_10_percentage': float(cv_lt_10_pct) * 100 if cv_lt_10_pct else 0,
            'cv_10_15_percentage': float(cv_10_15_pct) * 100 if cv_10_15_pct else 0,
            'cv_15_20_percentage': float(cv_15_20_pct) * 100 if cv_15_20_pct else 0,
            'cv_gt_20_percentage': float(cv_gt_20_pct) * 100 if cv_gt_20_pct else 0,
            'quality_rating': quality_rating,
            'test_type': 'serology',
            'marker_name': marker_name,
            'manufacturer_name': manufacturer_name,
            'inclusion_group': 'serology_extended',
        }

        configs_to_insert.append(config)

    wb.close()

    print(f'\nFound {len(configs_to_insert)} SEROLOGY configurations to import')
    print(f'Skipped {skipped} duplicates (already in "SEROLOGY to include")')

    if configs_to_insert:
        # Insert into database
        cursor = conn.cursor()

        insert_query = """
            INSERT INTO test_configurations (
                assay_name, qc_sample_name, events_examined,
                cv_lt_10_percentage, cv_10_15_percentage, cv_15_20_percentage, cv_gt_20_percentage,
                quality_rating, test_type, marker_name, manufacturer_name, inclusion_group
            ) VALUES %s
        """

        values = [(
            c['assay_name'], c['qc_sample_name'], c['events_examined'],
            c['cv_lt_10_percentage'], c['cv_10_15_percentage'], c['cv_15_20_percentage'], c['cv_gt_20_percentage'],
            c['quality_rating'], c['test_type'], c['marker_name'], c['manufacturer_name'], c['inclusion_group']
        ) for c in configs_to_insert]

        execute_values(cursor, insert_query, values)
        conn.commit()
        cursor.close()

        print(f'✓ Successfully imported {len(configs_to_insert)} SEROLOGY configurations')

    return len(configs_to_insert)

def main():
    print('='*80)
    print('RAW DATA IMPORT SCRIPT')
    print('='*80)
    print(f'\nExcel file: {EXCEL_PATH}')
    print(f'Database: {DATABASE_URL[:50]}...')

    # Connect to database
    conn = connect_db()

    try:
        # Import NAT data
        nat_count = import_nat_raw_data(conn)

        # Import SEROLOGY raw data
        serology_count = import_serology_raw_data(conn)

        # Summary
        print('\n' + '='*80)
        print('IMPORT COMPLETE')
        print('='*80)
        print(f'\nNAT data: {nat_count} rows imported')
        print(f'SEROLOGY raw data: {serology_count} rows imported')
        print(f'Total new rows: {nat_count + serology_count}')

        # Show final counts by inclusion_group
        cursor = conn.cursor()
        cursor.execute("""
            SELECT inclusion_group, test_type, COUNT(*) as count
            FROM test_configurations
            GROUP BY inclusion_group, test_type
            ORDER BY inclusion_group, test_type
        """)

        print('\nFinal database counts:')
        for row in cursor.fetchall():
            print(f'  {row[0]:25s} | {row[1]:10s} | {row[2]:4d} rows')

        cursor.close()

    except Exception as e:
        print(f'\nError during import: {e}')
        conn.rollback()
        sys.exit(1)
    finally:
        conn.close()

if __name__ == '__main__':
    main()
