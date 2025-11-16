#!/usr/bin/env python3
"""
Fast Raw Data Import - Simplified approach
Uses caching to minimize database queries
"""

import os
import sys
from openpyxl import load_workbook
import psycopg2
from dotenv import load_dotenv

load_dotenv('.env.local')

DATABASE_URL = os.getenv('DATABASE_URL')
EXCEL_PATH = 'build-data/documentation/Copy of 1. Summary Tables CV Events.xlsx'

# Caches for entity IDs
CACHE = {
    'categories': {},
    'pathogens': {},
    'markers': {},
    'manufacturers': {},
    'assays': {},
    'qc_samples': {},
}

def load_caches(conn):
    """Load existing entities into cache"""
    print('Loading existing entities into cache...')
    cursor = conn.cursor()

    # Categories
    cursor.execute("SELECT id, name FROM categories")
    for row in cursor.fetchall():
        CACHE['categories'][row[1]] = row[0]

    # Pathogens
    cursor.execute("SELECT id, name FROM pathogens")
    for row in cursor.fetchall():
        CACHE['pathogens'][row[1]] = row[0]

    # Markers
    cursor.execute("SELECT id, name FROM markers")
    for row in cursor.fetchall():
        CACHE['markers'][row[1]] = row[0]

    # Manufacturers
    cursor.execute("SELECT id, name FROM manufacturers")
    for row in cursor.fetchall():
        CACHE['manufacturers'][row[1]] = row[0]

    # Assays
    cursor.execute("SELECT id, name, manufacturer_id FROM assays")
    for row in cursor.fetchall():
        key = f"{row[1]}|{row[2]}"
        CACHE['assays'][key] = row[0]

    # QC Samples
    cursor.execute("SELECT id, name FROM qc_samples")
    for row in cursor.fetchall():
        CACHE['qc_samples'][row[1]] = row[0]

    cursor.close()
    print(f'  Cached: {len(CACHE["markers"])} markers, {len(CACHE["assays"])} assays, {len(CACHE["qc_samples"])} samples')

def get_or_create_simple(conn, table, name):
    """Get or create entity by name (simple tables)"""
    if name in CACHE[table]:
        return CACHE[table][name]

    cursor = conn.cursor()
    cursor.execute(f"INSERT INTO {table} (name) VALUES (%s) RETURNING id", (name,))
    entity_id = cursor.fetchone()[0]
    cursor.close()

    CACHE[table][name] = entity_id
    return entity_id

def get_or_create_marker(conn, name):
    """Get or create marker (simplified)"""
    if name in CACHE['markers']:
        return CACHE['markers'][name]

    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO markers (name, marker_type) VALUES (%s, %s) RETURNING id",
        (name, 'Antibody' if 'anti-' in name else 'Molecular')
    )
    marker_id = cursor.fetchone()[0]
    cursor.close()

    CACHE['markers'][name] = marker_id
    return marker_id

def get_or_create_assay(conn, name, manufacturer_name):
    """Get or create assay"""
    # Get manufacturer ID
    if manufacturer_name not in CACHE['manufacturers']:
        CACHE['manufacturers'][manufacturer_name] = get_or_create_simple(conn, 'manufacturers', manufacturer_name)

    manufacturer_id = CACHE['manufacturers'][manufacturer_name]
    key = f"{name}|{manufacturer_id}"

    if key in CACHE['assays']:
        return CACHE['assays'][key]

    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO assays (name, manufacturer_id) VALUES (%s, %s) RETURNING id",
        (name, manufacturer_id)
    )
    assay_id = cursor.fetchone()[0]
    cursor.close()

    CACHE['assays'][key] = assay_id
    return assay_id

def extract_manufacturer(assay_name):
    """Extract manufacturer from assay name"""
    assay_upper = assay_name.upper()
    manufacturers = {
        'ABBOTT': 'Abbott', 'ROCHE': 'Roche', 'SIEMENS': 'Siemens',
        'DIASORIN': 'DiaSorin', 'BIOMERIEUX': 'bioMerieux', 'BECKMAN': 'Beckman Coulter',
        'QIAGEN': 'QIAGEN', 'GRIFOLS': 'Grifols', 'ORTHO': 'Ortho Clinical Diagnostics',
        'LIAISON': 'DiaSorin', 'VIDAS': 'bioMerieux', 'ARCHITECT': 'Abbott',
        'ALINITY': 'Abbott', 'COBAS': 'Roche', 'ELECSYS': 'Roche',
        'ADVIA': 'Siemens', 'CENTAUR': 'Siemens', 'VITROS': 'Ortho Clinical Diagnostics',
        'AUSDIAGNOSTICS': 'AusDiagnostics', 'CEPHEID': 'Cepheid', 'HOLOGIC': 'Hologic',
    }
    for key, mfr in manufacturers.items():
        if key in assay_upper:
            return mfr
    return 'Unknown'

def parse_nat_marker(sample_name, assay_name):
    """Parse NAT marker name"""
    combined = (sample_name + ' ' + assay_name).upper()
    if 'CMV' in combined:
        return 'CMV'
    elif 'HCV' in combined:
        return 'HCV'
    elif 'HBV' in combined:
        return 'HBV'
    elif 'HIV' in combined:
        return 'HIV'
    elif 'HPV' in combined:
        return 'HPV'
    elif 'COVID' in combined or 'SARS' in combined:
        return 'SARS-CoV-2'
    elif 'CT' in combined and 'CHLAMYDIA' not in combined:
        return 'Chlamydia trachomatis'
    elif 'NG' in combined:
        return 'Neisseria gonorrhoeae'
    return 'Unknown NAT'

def parse_serology_marker(assay_name):
    """Parse serology marker name"""
    au = assay_name.upper()
    if 'CMV' in au and 'IGG' in au:
        return 'anti-CMV IgG'
    elif 'CMV' in au and 'IGM' in au:
        return 'anti-CMV IgM'
    elif 'EBV' in au and 'IGG' in au:
        return 'anti-EBV IgG'
    elif 'EBV' in au and 'IGM' in au:
        return 'anti-EBV IgM'
    elif 'TOXO' in au and 'IGG' in au:
        return 'anti-Toxoplasma IgG'
    elif 'TOXO' in au and 'IGM' in au:
        return 'anti-Toxoplasma IgM'
    elif 'RUBELLA' in au and 'IGG' in au:
        return 'anti-Rubella IgG'
    elif 'RUBELLA' in au and 'IGM' in au:
        return 'anti-Rubella IgM'
    elif 'HCV' in au:
        return 'anti-HCV'
    elif 'HAV' in au and 'IGM' in au:
        return 'anti-HAV IgM'
    elif 'HAV' in au and 'IGG' in au:
        return 'anti-HAV IgG'
    elif 'HAV' in au:
        return 'anti-HAV'
    elif 'HBSAG' in au:
        return 'HBsAg'
    elif 'HBS' in au:
        return 'anti-HBs'
    elif 'HBC' in au and 'IGM' in au:
        return 'anti-HBc IgM'
    elif 'HBC' in au:
        return 'anti-HBc'
    elif 'HBE' in au:
        return 'anti-HBe'
    elif 'HIV' in au and 'P24' in au:
        return 'HIV p24 antigen'
    elif 'HIV' in au and ('HIV-2' in au or 'HIV2' in au):
        return 'anti-HIV-2'
    elif 'HIV' in au:
        return 'anti-HIV-1+2'
    elif 'SYPHILIS' in au or 'TP' in au:
        return 'anti-Treponema pallidum'
    return 'Unknown'

def quality_rating(cv_pct):
    """Calculate quality rating"""
    if cv_pct >= 0.95:
        return 'excellent'
    elif cv_pct >= 0.85:
        return 'good'
    elif cv_pct >= 0.70:
        return 'acceptable'
    return 'poor'

def import_nat(conn):
    """Import NAT data"""
    print('\n' + '='*60)
    print('NAT DATA IMPORT')
    print('='*60)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['NAT raw data']

    imported = 0
    skipped = 0

    for row_idx in range(4, ws.max_row + 1):
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events = ws.cell(row_idx, 3).value
        cv_lt_10 = ws.cell(row_idx, 5).value or 0

        if not assay_name or not sample_name or assay_name == 'Assay':
            continue

        try:
            # Get IDs
            marker_name = parse_nat_marker(sample_name, assay_name)
            marker_id = get_or_create_marker(conn, marker_name)

            mfr_name = extract_manufacturer(assay_name)
            assay_id = get_or_create_assay(conn, assay_name, mfr_name)

            if sample_name not in CACHE['qc_samples']:
                CACHE['qc_samples'][sample_name] = get_or_create_simple(conn, 'qc_samples', sample_name)
            qc_sample_id = CACHE['qc_samples'][sample_name]

            # Insert config
            cursor = conn.cursor()
            cursor.execute(
                """INSERT INTO test_configurations (
                    marker_id, assay_id, qc_sample_id, test_type,
                    events_examined, quality_rating, inclusion_group
                ) VALUES (%s, %s, %s, 'nat', %s, %s, 'NAT raw data.csv')
                ON CONFLICT (marker_id, assay_id, qc_sample_id) DO NOTHING""",
                (marker_id, assay_id, qc_sample_id, int(events) if events else 0, quality_rating(cv_lt_10))
            )

            if cursor.rowcount > 0:
                imported += 1
            else:
                skipped += 1

            cursor.close()

            if (imported + skipped) % 10 == 0:
                print(f'  Processed: {imported + skipped} rows...', end='\r')
                conn.commit()  # Commit batch

        except Exception as e:
            print(f'\n  Error row {row_idx}: {e}')
            conn.rollback()

    conn.commit()  # Final commit
    wb.close()
    print(f'\n✓ NAT: {imported} imported, {skipped} skipped        ')
    return imported

def import_serology(conn):
    """Import SEROLOGY extended data"""
    print('\n' + '='*60)
    print('SEROLOGY EXTENDED IMPORT')
    print('='*60)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SEROLOGY raw data']

    imported = 0
    skipped = 0

    for row_idx in range(4, ws.max_row + 1):
        assay_name = ws.cell(row_idx, 1).value
        sample_name = ws.cell(row_idx, 2).value
        events = ws.cell(row_idx, 3).value
        cv_lt_10 = ws.cell(row_idx, 5).value or 0

        if not assay_name or not sample_name or assay_name == 'Assay':
            continue

        try:
            # Get IDs
            marker_name = parse_serology_marker(assay_name)
            marker_id = get_or_create_marker(conn, marker_name)

            mfr_name = extract_manufacturer(assay_name)
            assay_id = get_or_create_assay(conn, assay_name, mfr_name)

            if sample_name not in CACHE['qc_samples']:
                CACHE['qc_samples'][sample_name] = get_or_create_simple(conn, 'qc_samples', sample_name)
            qc_sample_id = CACHE['qc_samples'][sample_name]

            # Insert config
            cursor = conn.cursor()
            cursor.execute(
                """INSERT INTO test_configurations (
                    marker_id, assay_id, qc_sample_id, test_type,
                    events_examined, quality_rating, inclusion_group
                ) VALUES (%s, %s, %s, 'serology', %s, %s, 'SEROLOGY raw data.csv')
                ON CONFLICT (marker_id, assay_id, qc_sample_id) DO NOTHING""",
                (marker_id, assay_id, qc_sample_id, int(events) if events else 0, quality_rating(cv_lt_10))
            )

            if cursor.rowcount > 0:
                imported += 1
            else:
                skipped += 1

            cursor.close()

            if (imported + skipped) % 10 == 0:
                print(f'  Processed: {imported + skipped} rows...', end='\r')
                conn.commit()  # Commit batch

        except Exception as e:
            print(f'\n  Error row {row_idx}: {e}')
            conn.rollback()

    conn.commit()  # Final commit
    wb.close()
    print(f'\n✓ SEROLOGY: {imported} imported, {skipped} skipped        ')
    return imported

def main():
    print('='*60)
    print('RAW DATA IMPORT - FAST VERSION')
    print('='*60)

    try:
        conn = psycopg2.connect(DATABASE_URL)
        load_caches(conn)

        nat_count = import_nat(conn)
        serology_count = import_serology(conn)

        # Summary
        print('\n' + '='*60)
        print('FINAL SUMMARY')
        print('='*60)

        cursor = conn.cursor()
        cursor.execute("""
            SELECT inclusion_group, COUNT(*)
            FROM test_configurations
            GROUP BY inclusion_group
            ORDER BY inclusion_group
        """)

        print('\nDatabase totals:')
        total = 0
        for row in cursor.fetchall():
            print(f'  {row[0]:25s}: {row[1]:4d} configs')
            total += row[1]
        print(f'  {"TOTAL":25s}: {total:4d} configs')

        cursor.close()
        conn.close()

        print('\n✓ Import complete!')

    except Exception as e:
        print(f'\n❌ Error: {e}')
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
